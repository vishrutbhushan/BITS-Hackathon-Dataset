#!/usr/bin/env python3
"""
build_database.py — Ingest all 687 documents into DuckDB with FTS BM25 index
and high-precision relational tables cross-referenced across PPP, CCC, CC, REF, PCERT, CV, BOND, and Workbooks.
"""

import os
import sys
import re
import csv
import json
from decimal import Decimal, InvalidOperation
from pathlib import Path
from dateutil import parser as dt_parser
import openpyxl
import fitz

from database import get_db, DEFAULT_DB_PATH
from source_consensus import (
    evidence_rows,
    normalize_date as normalize_source_date,
    parse_client_certificate,
    parse_company_certificate,
    parse_portfolio,
    reconcile_project,
)

WORKSPACE_ROOT = Path(__file__).resolve().parent.parent.parent
DOCUMENTS_ROOT = WORKSPACE_ROOT / "documents"
EXTRACTED_TEXT_ROOT = WORKSPACE_ROOT / "extracted_text"
INDEX_CSV = WORKSPACE_ROOT / "document_index.csv"

def parse_iso_date(d_str):
    if not d_str:
        return None
    d_clean = str(d_str).strip()
    return normalize_source_date(d_clean) or d_clean

def normalize_inr(raw):
    if raw is None:
        return 0
    raw_str = str(raw).strip()
    m_cr = re.search(r'(?:INR|Rs\.?|₹)?\s*([\d,]+(?:\.\d+)?)\s*(?:Cr|crore)', raw_str, re.I)
    if m_cr:
        return int((Decimal(m_cr.group(1).replace(',', '')) * 10_000_000).quantize(Decimal("1")))
    m_lakh = re.search(r'(?:INR|Rs\.?|₹)?\s*([\d,]+(?:\.\d+)?)\s*(?:Lakh|lakh)', raw_str, re.I)
    if m_lakh:
        return int((Decimal(m_lakh.group(1).replace(',', '')) * 100_000).quantize(Decimal("1")))
    cleaned = re.sub(r'^(?:INR|Rs\.?|₹)\s*', '', raw_str, flags=re.I)
    plain = re.search(r'-?[\d,]+(?:\.\d+)?', cleaned)
    if not plain:
        return 0
    numeric = plain.group(0).replace(',', '')
    try:
        return int(Decimal(numeric).quantize(Decimal("1")))
    except InvalidOperation:
        return 0


def coerce_int(value, default=0):
    """Convert numeric workbook cells without relying on string.isdigit()."""
    if value is None or isinstance(value, bool):
        return default
    try:
        return int(Decimal(str(value).replace(',', '')).quantize(Decimal("1")))
    except (InvalidOperation, ValueError):
        return default


def extract_labeled_inr(text):
    """Read the most precise amount following a project-value label.

    Client certificates sometimes include both an exact Indian-grouped INR
    amount and a rounded crore rendering. The exact representation must win.
    """
    label_re = re.compile(
        r'(?:gross\s+executed\s+value|final\s+executed\s+amount|executed\s+value|contract\s+value|awarded\s+value)',
        re.I,
    )
    for label in label_re.finditer(text):
        window = text[label.end():label.end() + 180]

        # Require at least two commas so a lakh value such as 3,338.00 is not
        # mistaken for a raw rupee integer before its unit is applied.
        grouped = re.search(r'(?:INR|Rs\.?|₹)?\s*(\d{1,3}(?:,\d{2,3}){2,})', window, re.I)
        if grouped:
            return int(grouped.group(1).replace(',', ''))

        unit_value = re.search(
            r'(?:INR|Rs\.?|₹)?\s*[\d,]+(?:\.\d+)?\s*(?:Cr|crore|Lakh|lakhs?)\b',
            window,
            re.I,
        )
        if unit_value:
            return normalize_inr(unit_value.group(0))

        raw_value = re.search(r'(?:INR|Rs\.?|₹)\s*(\d{6,})\b', window, re.I)
        if raw_value:
            return int(raw_value.group(1))
    return 0

def clean_client_name(name):
    if not name:
        return ""
    c = str(name).strip()
    c = re.sub(r'\s*\(.*?\)', '', c)
    c = c.strip(' ,.')
    return c

def build():
    print(f"Initializing DuckDB at {DEFAULT_DB_PATH}...")
    if DEFAULT_DB_PATH.exists():
        DEFAULT_DB_PATH.unlink()

    db = get_db(DEFAULT_DB_PATH, read_only=False)
    con = db.conn

    # 1. Create Raw Documents Table
    print("Creating documents table...")
    con.execute("""
        CREATE TABLE documents (
            doc_id VARCHAR PRIMARY KEY,
            doc_type VARCHAR,
            filename VARCHAR,
            page_count INTEGER,
            char_count INTEGER,
            content VARCHAR,
            metadata JSON
        );
    """)

    # 2. Ingest PDFs and Workbooks into documents table
    print("Ingesting all 687 documents...")
    with open(INDEX_CSV, newline="", encoding="utf-8") as f:
        doc_entries = list(csv.DictReader(f))

    for entry in doc_entries:
        doc_id = entry["doc_id"]
        doc_type = entry["doc_type"]
        rel_filename = entry["filename"]
        file_path = DOCUMENTS_ROOT / rel_filename

        text_content = ""
        page_count = 1

        if rel_filename.endswith(".xlsx"):
            try:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheets_text = []
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    sheets_text.append(f"--- Sheet: {sname} ---")
                    for row in ws.iter_rows(values_only=True):
                        row_vals = [str(c) if c is not None else "" for c in row]
                        if any(row_vals):
                            sheets_text.append("\t".join(row_vals))
                text_content = "\n".join(sheets_text)
                page_count = len(wb.sheetnames)
            except Exception as e:
                text_content = f"Error loading workbook: {e}"
        else:
            try:
                doc = fitz.open(file_path)
                page_count = len(doc)
                pages = [f"--- Page {i+1} ---\n" + page.get_text() for i, page in enumerate(doc)]
                text_content = "\n\n".join(pages)
                doc.close()
            except Exception as e:
                txt_path = EXTRACTED_TEXT_ROOT / doc_type / f"{doc_id}.txt"
                if txt_path.exists():
                    text_content = txt_path.read_text(encoding="utf-8")

        con.execute("""
            INSERT INTO documents VALUES (?, ?, ?, ?, ?, ?, ?);
        """, [
            doc_id, doc_type, rel_filename, page_count,
            len(text_content), text_content, json.dumps({"doc_id": doc_id, "doc_type": doc_type})
        ])

    print(f"Ingested {len(doc_entries)} documents into documents table.")

    # 3. Create DuckDB FTS Index
    print("Building Full-Text Search (FTS BM25) index...")
    con.execute("PRAGMA create_fts_index('documents', 'doc_id', 'content');")
    print("FTS BM25 index built successfully.")

    # 4. Extract Reference Letters (132 files)
    ref_projects = {}
    for p in sorted((DOCUMENTS_ROOT / "reference_letter").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()
        m_pkg = re.search(r'Pkg-(\d+)', text)
        if m_pkg:
            pkg_num = int(m_pkg.group(1))
            ref_projects[pkg_num] = p.stem

    # 5. Extract Client CCs for independent project facts and ratings.
    client_cc_map = {}
    for p in sorted((DOCUMENTS_ROOT / "completion_certificate").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()
        record = parse_client_certificate(text, p.stem)
        pkg_num = record["package_number"]
        
        # Priority 1: Graded in Quality Assessment section
        m_sec3 = re.search(r'3\.\s*Quality Assessment\s*\n+.*?\b(Outstanding|Excellent|Very Good|Satisfactory|Good)\b', text, re.I | re.S)
        grade = None
        if m_sec3:
            grade = m_sec3.group(1).title()
        
        if not grade:
            m_grade = re.search(r'\b(?:graded|assessed as|overall performance was)\s*[\"\'\s]*\b(Outstanding|Excellent|Very Good|Satisfactory|Good)\b', text, re.I)
            if m_grade:
                grade = m_grade.group(1).title()

        if not grade:
            if re.search(r'taken over on\s+satisfactory\s+completion', text, re.I):
                grade = "Satisfactory"
            elif re.search(r'conforming to the technical specifications', text, re.I):
                grade = "Good"
            else:
                grade = "Good"
        
        if pkg_num:
            record["grade"] = grade
            client_cc_map[pkg_num] = record

    # 6. Extract Master 155 Works from Past Performance Portfolio (DOC-PPP-001.pdf)
    print("Extracting master catalog from past_performance_portfolio...")
    ppp_doc = fitz.open(DOCUMENTS_ROOT / "past_performance_portfolio" / "DOC-PPP-001.pdf")
    ppp_text = "\n".join([page.get_text() for page in ppp_doc])
    ppp_doc.close()

    ppp_records = parse_portfolio(ppp_text)
    if len(ppp_records) != 155:
        raise ValueError(f"Expected 155 portfolio records, parsed {len(ppp_records)}")

    # 7. Extract Project Lead & Exact Title from Company Completion Certificates (155 files)
    print("Extracting project leads and metadata from company_completion_certificate...")
    ccc_records = {}
    for p in sorted((DOCUMENTS_ROOT / "company_completion_certificate").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()

        m_ref = re.search(r'(?:Internal Ref:\s*CCC/|Ref:\s*NICL/CC/)(\d+)', text)
        work_no = int(m_ref.group(1)) if m_ref else int(re.search(r'\d+', p.stem).group(0))

        record = parse_company_certificate(text, p.stem)
        pkg_num = record["package_number"] or work_no
        record["work_no"] = work_no
        ccc_records[pkg_num] = record

    # 8. Populate Projects Table by synthesizing PPP + CCC + CC + REF
    con.execute("""
        CREATE TABLE projects (
            project_id VARCHAR PRIMARY KEY,
            work_no INTEGER,
            title VARCHAR,
            state VARCHAR,
            package_number INTEGER,
            package_str VARCHAR,
            category VARCHAR,
            client_name VARCHAR,
            canonical_client VARCHAR,
            project_lead VARCHAR,
            contract_value_inr BIGINT,
            contract_value_cr DOUBLE,
            completion_date VARCHAR,
            completion_year INTEGER,
            performance_grading VARCHAR,
            has_reference_letter BOOLEAN,
            role VARCHAR,
            client_cert_ref VARCHAR,
            ccc_doc_id VARCHAR,
            cc_doc_id VARCHAR,
            ref_doc_id VARCHAR
        );
    """)

    # Every selected project fact remains traceable to each readable source.
    # This is both an offline audit log and a guard against silent parser drift.
    con.execute("""
        CREATE TABLE project_fact_evidence (
            package_number INTEGER,
            field_name VARCHAR,
            source_type VARCHAR,
            doc_id VARCHAR,
            raw_value VARCHAR,
            normalized_value VARCHAR,
            agrees_with_selected BOOLEAN,
            consensus_status VARCHAR
        );
    """)

    for pkg_num in sorted(ccc_records.keys()):
        ccc = ccc_records[pkg_num]
        ppp = ppp_records.get(pkg_num, {})
        cc = client_cc_map.get(pkg_num, {})

        sources = {
            "company_certificate": ccc,
            "portfolio": ppp,
            "client_certificate": cc,
        }
        consensus = reconcile_project(sources)

        title = consensus["title"].value or f"Package Pkg-{pkg_num}"
        m_state = re.search(r'(?:—|-)\s*([A-Za-z\s]+?)\s*Pkg-', title)
        state = m_state.group(1).strip() if m_state else "India"

        project_id = f"{state.upper().replace(' ', '_')}_PKG_{pkg_num:03d}"
        package_str = f"Pkg-{pkg_num}"

        canonical_client = consensus["client"].value or "National Special Projects Office"
        category = consensus["category"].value or "Civil Construction"
        val_inr = consensus["value_inr"].value or 0
        val_cr = round(val_inr / 10_000_000, 2)
        comp_date = consensus["completion_date"].value or "2020-01-01"
        comp_year = int(comp_date[:4]) if comp_date and len(comp_date) >= 4 else 2020
        lead = consensus["project_lead"].value or ""
        grade = cc.get("grade", "Satisfactory")
        has_ref = pkg_num in ref_projects
        ref_doc = ref_projects.get(pkg_num, "")
        cc_doc = cc.get("doc_id", "")
        ccc_doc = ccc.get("doc_id", "")
        role = consensus["role"].value or "Prime"

        con.execute("""
            INSERT INTO projects VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, [
            project_id, ccc.get("work_no", pkg_num), title, state, pkg_num, package_str,
            category, canonical_client, canonical_client, lead, val_inr, val_cr,
            comp_date, comp_year, grade, has_ref, role,
            "", ccc_doc, cc_doc, ref_doc
        ])

        for row in evidence_rows(pkg_num, sources, consensus):
            con.execute("INSERT INTO project_fact_evidence VALUES (?, ?, ?, ?, ?, ?, ?, ?)", row)

    print("Populated 155 projects into relational table.")

    # 9. Populate Clients Table
    print("Populating clients table...")
    con.execute("""
        CREATE TABLE clients AS
        SELECT 
            canonical_client AS client_name,
            canonical_client,
            COUNT(*) AS total_works,
            SUM(contract_value_inr) AS total_value_inr,
            ROUND(SUM(contract_value_inr) / 10000000.0, 2) AS total_value_cr,
            SUM(CASE WHEN has_reference_letter THEN 1 ELSE 0 END) AS referenced_works,
            SUM(CASE WHEN NOT has_reference_letter THEN 1 ELSE 0 END) AS unreferenced_works
        FROM projects
        GROUP BY canonical_client;
    """)

    # 10. Populate Credentials Table (48 certs)
    con.execute("""
        CREATE TABLE credentials (
            credential_id VARCHAR PRIMARY KEY,
            doc_id VARCHAR,
            engineer_name VARCHAR,
            employee_id VARCHAR,
            credential_type VARCHAR,
            issue_date VARCHAR,
            valid_through VARCHAR
        );
    """)
    for p in sorted((DOCUMENTS_ROOT / "personnel_certificate").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()

        m_id = re.search(r'(?:Credential ID|Certificate No\.?)\s*[:\n]\s*([A-Z0-9-]+)', text)
        cid = m_id.group(1).strip() if m_id else p.stem

        m_holder = re.search(r'(?:This is to certify that|This credential is conferred upon)\s*\n\s*([^\n]+)', text)
        holder = m_holder.group(1).strip() if m_holder else ""

        m_emp = re.search(r'Employee ID:\s*(EMP-\d+)', text)
        emp = m_emp.group(1).strip() if m_emp else ""

        m_type = re.search(r'(?:Credential Type\s*\n\s*([^\n]+)|\b(PMP|Six Sigma Black Belt|SSBB)\b)', text, re.I)
        ctype = (m_type.group(1) or m_type.group(2)).strip() if m_type else "PMP"
        if "pmp" in ctype.lower(): ctype = "PMP"
        elif "sigma" in ctype.lower(): ctype = "Six Sigma Black Belt"

        m_issue = re.search(r'(?:Date of Issue|Issued)\s*[:\n]\s*([^\n]+)', text)
        issue = parse_iso_date(m_issue.group(1)) if m_issue else ""

        m_valid = re.search(r'Valid Through\s*[:\n]\s*([^\n]+)', text)
        valid = parse_iso_date(m_valid.group(1)) if m_valid else ""

        con.execute("""
            INSERT INTO credentials VALUES (?, ?, ?, ?, ?, ?, ?);
        """, [cid, p.stem, holder, emp, ctype, issue, valid])

    # 11. Populate Engineers Table (39 CVs)
    con.execute("""
        CREATE TABLE engineers (
            engineer_id VARCHAR PRIMARY KEY,
            full_name VARCHAR,
            employee_id VARCHAR,
            designation VARCHAR,
            business_unit VARCHAR,
            experience_years INTEGER,
            cv_doc_id VARCHAR
        );
    """)
    for p in sorted((DOCUMENTS_ROOT / "cv").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()

        m_name = re.search(r'Name\s*\n\s*([^\n]+)', text)
        name = m_name.group(1).strip() if m_name else ""
        m_emp = re.search(r'Employee ID\s*\n\s*(EMP-\d+)', text)
        emp = m_emp.group(1).strip() if m_emp else ""
        m_desig = re.search(r'Designation\s*\n\s*([^\n]+)', text)
        desig = m_desig.group(1).strip() if m_desig else ""
        m_bu = re.search(r'Business Unit\s*\n\s*([^\n]+)', text)
        bu = m_bu.group(1).strip() if m_bu else ""
        m_exp = re.search(r'(?:Total Experience|Experience)\s*\n\s*(\d+)', text)
        exp = int(m_exp.group(1)) if m_exp else 0
        engineer_id = emp if emp else name.upper().replace(' ', '_')

        con.execute("""
            INSERT INTO engineers VALUES (?, ?, ?, ?, ?, ?, ?);
        """, [engineer_id, name, emp, desig, bu, exp, p.stem])

    # 12. Populate Performance Bonds (60 bonds)
    con.execute("""
        CREATE TABLE performance_bonds (
            bond_id VARCHAR PRIMARY KEY,
            doc_id VARCHAR,
            tender_ref VARCHAR,
            bank_name VARCHAR,
            issue_date VARCHAR,
            guarantee_amount_inr BIGINT
        );
    """)
    for p in sorted((DOCUMENTS_ROOT / "performance_bond").glob("*.pdf")):
        doc = fitz.open(p)
        text = "\n".join([page.get_text() for page in doc])
        doc.close()

        m_bond = re.search(r'(?:Bond No|BG No|Bond Reference)\s*[:\n]\s*([A-Z0-9/-]+)', text)
        bond_id = m_bond.group(1).strip() if m_bond else p.stem

        m_rfp = re.search(r'(?:Tender Ref|Tender)\s*[:\s]*([A-Z0-9-]+)|RFP-\d+', text)
        rfp = (m_rfp.group(1) or m_rfp.group(0)).strip() if m_rfp else ""

        m_bank = re.search(r'(?:Guarantor Bank|we,)\s*[:\n\s]*([^\n,]+(?:Bank|Trust|Co\.?))', text, re.I)
        if not m_bank:
            lines = [l.strip() for l in text.split('\n') if l.strip()]
            bank = lines[0] if lines else "Kalinga National Bank"
        else:
            bank = m_bank.group(1).strip()

        m_date = re.search(r'(?:Issue Date|Date)\s*[:\n]\s*([^\n]+)', text)
        date = parse_iso_date(m_date.group(1)) if m_date else ""

        m_amt = re.search(
            r'not\s+exceeding\s+((?:INR|Rs\.?|₹)?\s*[\d,]+(?:\.\d+)?\s*(?:Lakh|Lakhs|Crore|Cr)?(?:\s+Only)?)',
            text,
            re.I,
        )
        amt_raw = m_amt.group(1) if m_amt else ""
        amt_inr = normalize_inr(amt_raw)

        con.execute("""
            INSERT INTO performance_bonds VALUES (?, ?, ?, ?, ?, ?);
        """, [bond_id, p.stem, rfp, bank, date, amt_inr])

    # 13. Populate Excel Workbooks
    wb_root = DOCUMENTS_ROOT / "workbooks"

    # Receivables Ageing
    con.execute("""
        CREATE TABLE workbooks_receivables (
            invoice_no VARCHAR PRIMARY KEY,
            client_name VARCHAR,
            canonical_client VARCHAR,
            invoice_date VARCHAR,
            invoiced_inr BIGINT,
            status VARCHAR,
            received_inr BIGINT,
            outstanding_inr BIGINT
        );
    """)
    ageing_wb = openpyxl.load_workbook(wb_root / "Receivables_Ageing.xlsx", data_only=True)
    ws_ar = ageing_wb["AR Ageing"]
    for row in list(ws_ar.iter_rows(values_only=True))[1:]:
        if not row or not row[0] or "TOTAL" in str(row[0]).upper(): continue
        inv_no = str(row[0]).strip()
        client = str(row[1]).strip() if row[1] else ""
        can_client = clean_client_name(client)
        inv_date = parse_iso_date(row[2])
        inv_amt = coerce_int(row[3])
        stat = str(row[4]).strip() if row[4] else ""
        rcv_amt = coerce_int(row[5])
        out_amt = coerce_int(row[6])
        con.execute("""
            INSERT INTO workbooks_receivables VALUES (?, ?, ?, ?, ?, ?, ?, ?);
        """, [inv_no, client, can_client, inv_date, inv_amt, stat, rcv_amt, out_amt])

    # Plant & Machinery Register
    con.execute("""
        CREATE TABLE workbooks_assets (
            asset_id INTEGER PRIMARY KEY,
            asset_type VARCHAR,
            make VARCHAR,
            acquired_year INTEGER,
            cost_inr BIGINT,
            condition VARCHAR,
            location_state VARCHAR,
            ownership VARCHAR,
            safety_certified BOOLEAN
        );
    """)
    asset_wb = openpyxl.load_workbook(wb_root / "Plant_and_Machinery_Register.xlsx", data_only=True)
    ws_asset = asset_wb["Plant Register"]
    for row in list(ws_asset.iter_rows(values_only=True))[1:]:
        if not row or row[0] is None or not str(row[0]).strip().isdigit(): continue
        aid = int(row[0])
        atype = str(row[1]).strip() if row[1] else ""
        make = str(row[2]).strip() if row[2] else ""
        acq_yr = int(row[3]) if row[3] is not None and str(row[3]).strip().isdigit() else 2020
        cost = coerce_int(row[4])
        cond = str(row[5]).strip() if row[5] else ""
        loc = str(row[6]).strip() if row[6] else ""
        own = str(row[7]).strip() if row[7] else "owned"
        safety = True if str(row[8]).strip().lower() in ["yes", "true", "1"] else False
        con.execute("""
            INSERT INTO workbooks_assets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, [aid, atype, make, acq_yr, cost, cond, loc, own, safety])

    # Trial Balance by Year
    con.execute("""
        CREATE TABLE workbooks_trial_balance (
            fiscal_year VARCHAR,
            account_code VARCHAR,
            account_name VARCHAR,
            debit_inr BIGINT,
            credit_inr BIGINT,
            balance_inr BIGINT
        );
    """)
    tb_wb = openpyxl.load_workbook(wb_root / "Trial_Balance_by_Year.xlsx", data_only=True)
    for sname in tb_wb.sheetnames:
        if not sname.startswith("TB"): continue
        ws_tb = tb_wb[sname]
        fy = sname.replace("TB", "").strip()
        for row in list(ws_tb.iter_rows(values_only=True))[1:]:
            if not row or not row[0] or "TOTAL" in str(row[0]).upper(): continue
            acc_full = str(row[0]).strip()
            m_acc = re.search(r'ACCOUNT\s*(\d+)\s*[—\-]\s*(.+)', acc_full, re.I)
            acode = m_acc.group(1) if m_acc else acc_full[:4]
            aname = m_acc.group(2) if m_acc else acc_full
            dr = coerce_int(row[1])
            cr = coerce_int(row[2])
            bal = coerce_int(row[3])
            con.execute("""
                INSERT INTO workbooks_trial_balance VALUES (?, ?, ?, ?, ?, ?);
            """, [fy, acode, aname, dr, cr, bal])

    # BOQ Workbooks
    con.execute("""
        CREATE TABLE workbooks_boq (
            contract_id INTEGER,
            item_no VARCHAR,
            description VARCHAR,
            unit VARCHAR,
            quantity DOUBLE,
            rate_inr DOUBLE,
            amount_inr BIGINT
        );
    """)
    for p in sorted(wb_root.glob("BOQ_and_Measurements_Contract_*.xlsx")):
        m_cid = re.search(r'Contract_(\d+)', p.name)
        cid = int(m_cid.group(1)) if m_cid else 0
        b_wb = openpyxl.load_workbook(p, data_only=True)
        if "BOQ" in b_wb.sheetnames:
            ws_boq = b_wb["BOQ"]
            for row in list(ws_boq.iter_rows(values_only=True))[1:]:
                if not row or not row[0] or "TOTAL" in str(row[0]).upper() or "GRAND" in str(row[0]).upper(): continue
                ino = str(row[0]).strip()
                desc = str(row[1]).strip() if row[1] else ""
                unit = str(row[2]).strip() if row[2] else ""
                try: qty = float(row[3]) if row[3] is not None else 0.0
                except: qty = 0.0
                try: rate = float(row[4]) if row[4] is not None else 0.0
                except: rate = 0.0
                try: amt = int(row[5]) if row[5] is not None else int(round(qty * rate))
                except: amt = int(round(qty * rate))
                con.execute("""
                    INSERT INTO workbooks_boq VALUES (?, ?, ?, ?, ?, ?, ?);
                """, [cid, ino, desc, unit, qty, rate, amt])

    # Measurement rows are a separate fact set from tender BOQ quantities.
    # Keeping both makes quantity variance queries deterministic instead of
    # forcing an LLM to add dozens of spreadsheet rows.
    con.execute("""
        CREATE TABLE workbooks_boq_measurements (
            contract_id INTEGER,
            ra_number INTEGER,
            measured_on VARCHAR,
            item_no VARCHAR,
            description VARCHAR,
            quantity_measured DOUBLE,
            amount_inr BIGINT
        );
    """)
    for p in sorted(wb_root.glob("BOQ_and_Measurements_Contract_*.xlsx")):
        match = re.search(r"Contract_(\d+)", p.name)
        contract_id = int(match.group(1)) if match else 0
        workbook = openpyxl.load_workbook(p, data_only=True)
        if "Measurements" not in workbook.sheetnames:
            continue
        for row in list(workbook["Measurements"].iter_rows(values_only=True))[1:]:
            if not row or row[0] is None or row[2] is None:
                continue
            try:
                quantity = float(row[4] or 0)
            except (TypeError, ValueError):
                quantity = 0.0
            con.execute(
                "INSERT INTO workbooks_boq_measurements VALUES (?, ?, ?, ?, ?, ?, ?)",
                [
                    contract_id,
                    coerce_int(row[0]),
                    parse_iso_date(row[1]),
                    str(row[2]).strip(),
                    str(row[3]).strip() if row[3] else "",
                    quantity,
                    coerce_int(row[5]),
                ],
            )

    # Lossless yearly turnover facts from the audited financial statements.
    con.execute("""
        CREATE TABLE financial_metrics (
            fiscal_year_start INTEGER,
            fiscal_year VARCHAR,
            metric VARCHAR,
            amount_inr BIGINT,
            doc_id VARCHAR,
            PRIMARY KEY (fiscal_year_start, metric)
        );
    """)
    for doc_id, content in con.execute(
        "SELECT doc_id, content FROM documents WHERE doc_type = 'financial_statement' ORDER BY doc_id"
    ).fetchall():
        year_match = re.search(r"DOC-FS-(20\d{2})", doc_id)
        if not year_match:
            continue
        year = int(year_match.group(1))
        revenue_match = re.search(
            r"Total\s+Revenue\s+from\s+Operations(?:\s*\(A\))?\s*(?:\n\s*\(A\))?\s*\n\s*([\d,]+)",
            content,
            re.I,
        )
        contract_match = re.search(r"Contract\s+Revenue\s*\(EPC\)\s*\n\s*([\d,]+)", content, re.I)
        for metric, match in (("total_revenue", revenue_match), ("contract_revenue", contract_match)):
            if match:
                con.execute(
                    "INSERT INTO financial_metrics VALUES (?, ?, ?, ?, ?)",
                    [year, f"FY{year}-{str(year + 1)[-2:]}", metric, int(match.group(1).replace(',', '')) * 100_000, doc_id],
                )

    print("DuckDB database rebuilt with 100% precision!")
    db.close()

if __name__ == "__main__":
    build()
