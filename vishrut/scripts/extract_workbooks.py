#!/usr/bin/env python3
"""
Stage 0b: extract every sheet of every xlsx workbook to plain
row-of-dicts JSON, with formulas EVALUATED (openpyxl data_only=True
reads the last-calculated value Excel stored, not the formula text --
if you need live re-evaluation, open once in real Excel/LibreOffice
first to force a recalculation and save, since openpyxl itself doesn't
execute formulas).

Usage:
    pip install openpyxl --break-system-packages
    python extract_workbooks.py --index document_index.csv --root /path/to/parent/of/documents --out workbooks_raw
"""
import argparse
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run: pip install openpyxl --break-system-packages")


def extract_workbook(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        out[sheet_name] = rows
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--index", required=True)
    ap.add_argument("--root", required=True)
    ap.add_argument("--out", default="workbooks_raw")
    args = ap.parse_args()

    root = Path(args.root)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    with open(args.index, newline="", encoding="utf-8") as f:
        entries = [r for r in csv.DictReader(f) if r["filename"].lower().endswith((".xlsx", ".xls"))]

    print(f"Extracting {len(entries)} workbooks...")
    for row in entries:
        doc_id, filename = row["doc_id"], row["filename"]
        wb_path = root / filename
        if not wb_path.exists():
            print(f"  [skip] {filename} not found")
            continue
        data = extract_workbook(wb_path)
        out_path = out_dir / f"{doc_id}.json"
        out_path.write_text(json.dumps(data, default=str, indent=2), encoding="utf-8")
        print(f"  {doc_id}: {list(data.keys())}")

    print(f"\nWorkbook data written under: {out_dir}/<doc_id>.json")


if __name__ == "__main__":
    main()
